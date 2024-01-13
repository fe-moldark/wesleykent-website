# Script to auto click through a series of screens for the audit of a security
# system, then parse through that csv data, manipulate it into a workable
# format, and re-export it
#
# Note that some file locations and other identifying info has been changed
#
# Author: Wesley Kent
# Date: 10/01/2024

import os, sys
import time
import csv
from datetime import datetime
import getpass
from openpyxl import load_workbook #openpyxl==3.1.2
import pyautogui #PyAutoGUI==0.9.54
from pynput.keyboard import Key, Controller #pynput==1.7.6
import pyperclip #pyperclip==1.8.2
import keyboard #keyboard==0.13.5
from win32api import GetKeyState
import string
import winsound
import colorama #colorama==0.4.6
from colorama import Fore, Style #colorama==0.4.6

"""
Other modules used:
auto-py-to-exe==2.42.0
py2exe==0.13.0.1
pyinstaller==6.3.0
pyinstaller-hooks-contrib==2023.11

Running: Python 3.11.7
"""


colorama.init(True)

def Logging(text_to_write):
    logit=open("Log/File/Location/combinedLogAndErrorFile.txt","a") #this was changed, of course
    logit.write(text_to_write)
    logit.close()

#audio stuff
duration = 500
freq = 550

current_time = datetime.now()
usedate = current_time.strftime('%b %d, %Y')
user=getpass.getuser()


Logging('\n\nNew Error / Event Log // User: '+str(user)+' // Date/time: @'+str(current_time)+':\n\n ---- Begin Program ----\n')

#CSVs directory needs to hold the documents
folder_loc='Folder/Holding/TheCSVFiles/CSVs/'
checkThese=os.listdir(folder_loc)

scroll_main=[colorama.Fore.BLUE + 'Welcome to the Automated ------- Audit Program.',f'{Fore.WHITE}\nWritten by: Wesley Kent, Dec 2023',f'{Fore.YELLOW}\nNOTE: display MUST be set to 1920x1080 and 125% zoom or this program WILL NOT function as designed.',f'{Fore.BLUE}\nThis section of the program will walk you through how to collect the Card data for users and will then automatically convert and save this data in a workable format.',f'{Fore.LIGHTRED_EX}\nEvent error will be saved to: Log/File/Location/combinedLogAndErrorFile.txt. This also functions as a general log file.']
scroll_directions=[f'{Fore.WHITE}\n1) These are baseline instructions to download data about security cards (badges) being used by our system.','2) You can adjust these steps as needed.']

def scroll_function(which_list,which_time):
    for item in which_list:
        print(item)
        time.sleep(which_time)

scroll_function(scroll_main,1.5)
time.sleep(.5)
scroll_function(scroll_directions,.3)

winsound.Beep(freq, duration) #beeping sounds will indicate when action is needed from the user
just_waiting=input(f"{Fore.YELLOW}3) Once [above steps] are done, press the [ENTER] key on this program.")



##############################################
#Begin code for the auto clicker program here
##############################################


def key_down(key): #used to monitor [ENTER] key later on
    state = GetKeyState(key)
    if (state != 0) and (state != 1):
        return True
    else:
        return False

#Okay chars to use for filenames
valid_chars = "-_.() %s%s" % (string.ascii_letters, string.digits)
keyboard = Controller()


print (f'{Fore.BLUE}\n\nThis section of the program will walk you through the automated clicker program.')
time.sleep(1)
print(f'{Fore.WHITE}\n1) To begin the program, first complete the download of the first Access Level manually, from start to finish. The file will need to be saved to the Folder/Holding/TheCSVFiles/CSVs/" directory.')
time.sleep(0.3)
"""
These were steps to complete the first access level manually, very simple
"""
winsound.Beep(freq, duration)
waitOnEnter=input(f'{Fore.YELLOW}4) When finished, press the [ENTER] key.')
time.sleep(.65)

#Base coordinates for first stage
AccessLevelMenuDropdown=(-691,120)
EnglishTextBox=(-421,104)
LShapedBox=(-946,37)
#Second stage/screen
SaveToCSVButton=(-27,226)
CloseLShapedBox=()


def getModifiers(mod_num):
    while not key_down(0x0D):
        asString=str(pyautogui.position())[8:]
    #print('pyautogui.position():',pyautogui.position())

    gatherX=''
    while asString[0] in [str(i) for i in range(10)]+['-']: #stupid negative numbers depending on screen positioning...
        gatherX+=asString[0]
        asString=asString[1:]

    asString=asString[4:]
    gatherY=''
    while asString[0] in [str(i) for i in range(10)]:
        gatherY+=asString[0]
        asString=asString[1:]

    modifier_x_returned=int(gatherX)#-int(should_be_x)
    modifier_y_returned=int(gatherY)#-int(should_be_y)

    Logging('\n * '+str(mod_num)+' Modifier used: '+str(modifier_x_returned)+", "+str(modifier_y_returned))
    print(f"{Fore.WHITE}\nThe "+str(mod_num)+" modifier from the base that will be used is x=",str(modifier_x_returned)," y=",str(modifier_y_returned),".")

    return modifier_x_returned,modifier_y_returned

print(f"{Fore.WHITE}\nNow, center the point of your mouse over the Red exit button in the top right corner of the ACCESS LEVEL screen. Once you have done so (and with the FOCUS still being on this program),"+f'{Fore.YELLOW} press the [ENTER] key once.')
winsound.Beep(freq, duration)

modifier_x,modifier_y=getModifiers("1st")
time.sleep(1)

print(f"{Fore.WHITE}\nSimilar to before, now click the Red 'L' button that brings up the second screen. Center the point of your mouse over the Red exit button in the top right corner of this screen. Once you have done so (and with the FOCUS still being on this program), "+f'{Fore.YELLOW} press the [ENTER] key once.')
winsound.Beep(freq, duration)

CloseLShapedBox=(getModifiers("2nd"))


def loop():

    #click twice at the access level dropdown menu
    pyautogui.moveTo(AccessLevelMenuDropdown[0]+modifier_x, AccessLevelMenuDropdown[1]+modifier_y, duration = .2)
    time.sleep(0.2)
    pyautogui.click()
    time.sleep(0.5)
    pyautogui.click()
    time.sleep(0.3)

    #down key (to select next access level)
    keyboard.press(Key.down)
    keyboard.release(Key.down)
    time.sleep(0.4)

    #move mouse to english box and click once
    pyautogui.moveTo(EnglishTextBox[0]+modifier_x, EnglishTextBox[1]+modifier_y, duration = .2)
    pyautogui.click()
    time.sleep(0.4)

    #CTRL+A, CTRL+C --> YES, the breaks are needed
    keyboard.press(Key.ctrl)
    time.sleep(0.2)
    keyboard.press('a')
    time.sleep(0.2)
    keyboard.release('a')
    time.sleep(0.2)
    keyboard.release(Key.ctrl)
    time.sleep(0.6)
    keyboard.press(Key.ctrl)
    time.sleep(0.2)
    keyboard.press('c')
    time.sleep(0.2)
    keyboard.release('c')
    time.sleep(0.2)
    keyboard.release(Key.ctrl)
    time.sleep(0.3)

    getname1=pyperclip.paste()
    time.sleep(0.1)

    #Move mouse to L box and click
    pyautogui.moveTo(LShapedBox[0]+modifier_x, LShapedBox[1]+modifier_y, duration = .2)
    pyautogui.click()
    time.sleep(0.4)

    #Move to csv and click
    pyautogui.moveTo(SaveToCSVButton[0]+CloseLShapedBox[0], SaveToCSVButton[1]+CloseLShapedBox[1], duration = .2)
    pyautogui.click()
    time.sleep(3)

    #See if name needs to be modified
    listed=[letter for letter in getname1 if letter in valid_chars]

    base_folder='Folder/Holding/TheCSVFiles/CSVs/CSVs/'

    baseNum=2
    if os.path.isfile(base_folder+str(''.join(listed))+'.csv'): #there are a few instances where files have the same name, and it just overwrites them without this
        listed.append(' ')
        listed.append(str(baseNum))
        while os.path.isfile(base_folder+str(''.join(listed))+'.csv'):
            baseNum+=1
            listed[-1]=str(baseNum) #there should never be more than two of the same filenames here

            if baseNum>=10: #caught in a loop somewhere, troubleshoot and log it
                Logging('\n ! ERROR: the file path number checks exceeded 10 for the following file: '+str(base_folder+str(''.join(listed))+'.csv')+'. This is likely a fault on the software side, not to do with the Access Levels.')

                print(f'{Fore.LIGHTRED_EX}You ran into an error that should not be happening - reference error log now at: Log/File/Location/combinedLogAndErrorFile.txt')
                waitingOnInput=input(f"{Fore.YELLOW}Press the [ENTER] key to exit the program now.")
                sys.exit()

    time.sleep(0.4)

    for item in listed: #never had an issue with needing sleep() functions for this part
        keyboard.press(str(item))
        keyboard.release(str(item))
    time.sleep(1)

    #Press enter to save
    keyboard.press(Key.enter)
    keyboard.release(Key.enter)
    time.sleep(0.4)

    #Press enter again to acknowledge
    keyboard.press(Key.enter)
    keyboard.release(Key.enter)
    time.sleep(0.4)

    #Mouse click to close out 'L' option
    pyautogui.moveTo(CloseLShapedBox[0], CloseLShapedBox[1], duration = .2)
    pyautogui.click()
    time.sleep(0.5)


time.sleep(3.5)

print(f'{Fore.YELLOW}\nNow exit out of that screen and return to the Access Level screen, leaving the focus there. You have 10 seconds.\n')

winsound.Beep(freq, int(100))
winsound.Beep(freq, int(100))
winsound.Beep(freq, int(100))

time.sleep(9)

#print(pyautogui.position())
start=time.time()
loop() #just once for the user to confirm before starting the next 60+ loops...

print(f"{Fore.WHITE}This should have download the first Access Level. If it did so correctly and you wish to continue with the rest of the Access Levels, enter 'Yes' now. Otherwise, enter 'No' to end this section of the program and complete the rest manually, fix the program, or simply skip this stage for whatever reason.")
continue_yesno='NOTANANSWER'
answers=['yes','Yes','y','Y','no','No','n','N']
while continue_yesno not in answers:
    continue_yesno=input(f'{Fore.YELLOW}Input now: ')

if continue_yesno in ['no','No','n','N']:
    print(f'{Fore.WHITE}You indicated no, so the rest of this program will continue without the auto clicker function.')
    time.sleep(2)
else:
    # continue onwards...
    print(f'{Fore.YELLOW}You have elected to continue. You have 10 seconds to click anywhere back on the Server workstation to focus on that screen.')
    winsound.Beep(freq, int(100))
    winsound.Beep(freq, int(100))
    winsound.Beep(freq, int(100))
    time.sleep(9)

    for i in range(60): #will become length of list (62)-2, 1 for manual, 1 for the check
        loop()

    end=time.time()
    print(f'{Fore.GREEN}\n\nThe auto clicker section is complete.')
    time.sleep(0.3)
    winsound.Beep(freq, int(100))
    winsound.Beep(freq, int(100))
    winsound.Beep(freq, int(100))
    print(f'{Fore.WHITE}Total time (in seconds) to execute script: ',str(end-start))

    print(f'{Fore.WHITE}\n\nThis process looped through the current 62 different Access Levels. If there have been any additions since the creation of this program, complete the rest of them manually now and modify the script for the next go around.')
    anything=input(f"{Fore.YELLOW}\n\nOnce ready to continue, press the [ENTER] key.")

    totalTime=str(end-start)
    convertTime=totalTime[:totalTime.index('.')]

    Logging('\n ++++ Successfully executed auto clicker script. Time elapsed (in minutes)'+str(int(convertTime)/60)+' ++++ ')

###########################################
#End code for the auto clicker program here
###########################################


cardNumbersOnly=[]

def convert_all():
    try:
        test_file=open('Folder/Holding/TheCSVFiles/CSVs/CSVs/result/Cards.csv','r')
        test_read=test_file.read().splitlines()
        test_file.close()

    except FileNotFoundError:
        Logging(' ! FileNotFoundError: Folder/Holding/TheCSVFiles/CSVs/CSVs/result/Cards.csv file not found.\n')

        print(f'{Fore.LIGHTRED_EX}\n\nFileNotFoundError - Reference error log at: Log/File/Location/combinedLogAndErrorFile.txt')
        anything=input(f"{Fore.YELLOW}Press [ENTER] to exit program.")
        sys.exit()

    master_list=[]
    for row in test_read:
        split_again=row.split(",") #Formatted as csv...

        UID=split_again[2]
        LastName=split_again[3]
        FirstName=split_again[4]

        MasterValidInvalid=str(split_again[-11]) #this records the user as a whole, so even 'valid' individual cards are INvalidated by this value

        #the security software here is HORRIBLE - they do not put some fields containing a comma in quotation marks so it can't be used until it is modified
        #So a file containing "LastName, FirstName" messes up that line
        if FirstName[0]==" ": #means there is a first AND last name, seperated by a comma
            Username=str(LastName)+","+str(FirstName)
            FirstName=FirstName[1:]
        else:
            Username=str(LastName) #LN only
            FirstName='' #I did a check, the only cases in this exception are #s, ranging from 0,1,4-->User card states, not a FirstName

        possible_card_numbers=[(-i*2+1,-i*2) for i in range(1,6)] #overcomplicated, but more fun this way

        addCard=[]

        for pair in possible_card_numbers: #-1 through -10, 5 pairs total
            if (split_again[pair[0]],split_again[pair[1]])!=("",""): #Card data to grab

                #check if card is valid first
                if int(split_again[pair[0]])==0 and MasterValidInvalid=="0": #Note: 0=valid, 1=invalid, 2=stolen/lost, 4=USER state is invalid (or something like that)
                    addCard.append([split_again[pair[0]],split_again[pair[1]]])
                    cardNumbersOnly.append(str(split_again[pair[1]]))#
                else: #ignore invalid / lost states for now
                    pass
        
        for item in addCard:
            master_list.append([str(item[1]),str(LastName),str(FirstName),str(Username),str(UID),str(item[0]),''])

        """ Data formatting notes here:
        split_again[0] = + (first column ignore)
        split_again[1] = Card # (ignore)
        split_again[2] = UID (this is sometimes blank, but still try and pull it)
        split_again[3] = Username (part1)
        split_again[4] = Username (part2)

        Note: split_again[3] and [4] are only two values when the Card user name is a "Last, First"
        Can identify these cases when there is a space as the first char in split_again[3]
        The software is horrible for not noticing this and wrapping the cell value in quotation marks like they should

        Regardless, the rest of the data is pulled 'backwards' to avoid this possible formatting error
        split_again[-1] = Card 5 state
        split_again[-2] = Card 5 card number
        split_again[-3] = Card 4 state
        split_again[-4] = Card 4 card number

        Rinse & repeat...
        """

    Logging('\n * Note: lines in test_read: '+str(len(test_read))+'\n')
    Logging(' * Note: lines saved to master_list: '+str(len(master_list))+'\n')

    return master_list
 

master_list=convert_all()


try:
    with open('File1/You/Are/Creating/copyToCardsTab.csv', 'w') as f:
        write = csv.writer(f)
        headers=['Card #','Last Name','First Name','Username','UID','Card State'] #Note for self: for now the Card State column only contains valid (ie=0) states
        write.writerow(headers)
        write.writerows(master_list)

    #fml, but this works
    def cleanup_csv(location):
        fix=open(location, 'r')
        lines=fix.read().splitlines()

        lines=[line for line in lines if len(line)!=0]
        fix.close()

        fixed=open(location, 'w')
        for line in lines:
            if line!=lines[-1]:
                fixed.writelines(line+"\n")
            else:
                fixed.writelines(line)
        fixed.close()

    cleanup_csv('Folder/Holding/TheCSVFiles/CSVs/result/copyToCardsTab.csv')

except PermissionError:
    Logging(' ! PermissionsError. You likely have File1/You/Are/Creating/copyToCardsTab.csv open in Excel or Notepad that is preventing this from being overwritten with new data. Close it, and try again.\n')

    print(f'{Fore.LIGHTRED_EX}\n\nPermissionError - Reference error log at: Log/File/Location/combinedLogAndErrorFile.txt')
    anything=input(f"{Fore.YELLOW}Press [ENTER] to exit program.")

    sys.exit()


Logging(' ++++ Cards tab script executed successfully ++++\n')

time.sleep(1.5)
print(f'{Fore.GREEN}\nComplete. The file has been saved to File1/You/Are/Creating/copyToCardsTab.csv, and the data there should be copied into the "Cards" sheet of the Audit spreadsheet.')
time.sleep(.3)

try:
    os.system('start EXCEL.EXE "File1/You/Are/Creating/copyToCardsTab.csv"')
except:
    print(f'{Fore.LIGHTRED_EX}It looks like Excel is not installed on this device - no matter, the rest of the script will continue.')
    time.sleep(1)
    Logging(' ! Error: Trying to open File1/You/Are/Creating/copyToCardsTab.csv with Excel.\n')


#Next part of the script - copy over to 'Audit Report' tab
time.sleep(1)
print(f"{Fore.BLUE}\n\nThe 'Audit Report' section of this program will now begin.\n\n")
time.sleep(1)


try:

    def getAccessLevels():

        base_file_loc='Base/Folder/With/File/Of/Sheet/Containg/AccessLevelDefinitions/'
        file_loc='Audit.xlsx'

        if os.path.isfile(base_file_loc+file_loc):
            pass
        else:
            Logging(' ! Error: Failed to open "Base/Folder/With/File/Of/Sheet/Containg/AccessLevelDefinitions/Audit.xlsx" - Failed isfile() check. User prompted for new file.\n')

            print(f"{Fore.WHITE}The file listed as: "+Fore.BLUE+str(file_loc)+f"{Fore.WHITE} does not exist, which is needed to pull Access Level Definitions. Enter an alternative filename below (must be in the 'Base/Folder/With/File/Of/Sheet/Containg/AccessLevelDefinitions/' directory), or enter 'exit' to exit the program.\n")
            while not os.path.isfile(base_file_loc+file_loc):
                file_loc=input(f"{Fore.YELLOW}Enter filename or 'exit': ")

                if file_loc in ['exit','Exit']:
                    sys.exit()

            print(f"{Fore.GREEN}An alternative file was chosen called: "+str(file_loc))

            Logging(' * An alternate Audit spreadsheet was chosen, located at: '+str(file_loc)+'\n')

        
        wb = load_workbook(base_file_loc+file_loc)
        ws = wb["Access Level Definitions"] #only sheet we care about from this file

        first_column = ws['A'] #wherever those values for you are stored, for me column A

        AccessLevels=[]

        for x in range(1,len(first_column)): 
            AccessLevels.append(str(first_column[x].value))

        return AccessLevels

    AccessLevels=getAccessLevels()


    #CSVs directory needs to hold the documents
    folder_loc='Folder/Holding/TheCSVFiles/CSVs/'
    checkThese=os.listdir(folder_loc)

    time.sleep(1)
    print(f'{Fore.BLUE}\nBeginning the next stage for Audit Report Tab.')
    time.sleep(1.5)
    print(f'{Fore.WHITE}\nOnce running, this program does not require interaction UNLESS there are new Access Levels that have been created.')
    time.sleep(5)

    print(f'{Fore.WHITE}\n\nBefore filtering for empty files: '+f'{Fore.LIGHTCYAN_EX}'+str(len(checkThese)))
    for maybenotafile in checkThese:
        if os.path.isfile(folder_loc+maybenotafile):
            #check if file includes keys to look through - some (many, actually) are empty
            quickly=open(folder_loc+maybenotafile,"r")
            quickly_read=quickly.read().splitlines()
            quickly.close()

            if len(quickly_read)<=2:#means file is empty except for the headers, so remove
                checkThese.remove(maybenotafile)
                Logging(' ~ Empty Access Level identified: "'+str(maybenotafile)+'". Consider removing the Access Level.')

        else:
            checkThese.remove(maybenotafile)

    #print('After filtering: ',len(checkThese))
    time.sleep(0.3)   
    print(f'{Fore.WHITE}\nAfter filtering: '+f'{Fore.LIGHTCYAN_EX}'+str(len(checkThese)))  
    time.sleep(0.3)   
    #Keep seperate for now, but only saving to one CSV moving forward
    GoodCopy=[] #AccessLevel name exists
    BadCopy=[]  #AccessLevel name does not exists, need to find alt or request input from user


    #for troublehsooting, log when an alt. name is found and used
    def logNewNameUsed(oldname,newname,type):
        Logging(' * '+str(type)+'Using alt. Access Level name of: "'+str(newname)+'" in place of: "'+str(oldname)+'"\n')
        

    #this function will use common alternate spellings for filenames and find their corresponding Access Level names
    def known_alternatives(lookfor):

        #this dict contains respellings of known 'bad' names - keep in mind the files that are saved may have to remove characters
        #to save it with a valid filename (some special characters will stop this)
        dict_alt={
            "bad name":"good name" #for me, this list had ~20 elements to it
        }

        if dict_alt.get(lookfor) is not None:
            new_name=dict_alt.get(lookfor)
        else:
            new_name=lookfor #returns the OG name
        
        return new_name

    #initial sorting of files
    for filename in checkThese:
        if str(filename[:-4]) in AccessLevels: #Validate filenames (minus extension) against the AccessList Sheet, column A
            GoodCopy.append(str(filename))
        else:
            BadCopy.append(str(filename))

    #print(f'{Fore.WHITE}\n\nBefore filtering for empty files: '+f'{Fore.LIGHTCYAN_EX}'+str(len(checkThese)))
    print(f'{Fore.WHITE}\nNaming issues with: '+f'{Fore.LIGHTCYAN_EX}'+str(len(BadCopy)))
    time.sleep(0.3)
    print(f'{Fore.WHITE}No naming issues with: '+f'{Fore.LIGHTCYAN_EX}'+str(len(GoodCopy)))
    time.sleep(0.3)
    print(f"{Fore.WHITE}If you need to assign an access level name but it has not been defined yet in the Audit.xlsx spreadsheet, you may use 'Pass' or 'Skip' to name them later.\n")
    time.sleep(0.3)
    arrayToCSV=[]
    output='copyToAuditReportTab.csv'

    #below values just for logging data
    countGood=0
    countBad=0

    #main grab card info script
    def grabCardInfo(whichList,AccessLevels,arrayToCSV,countGood,countBad): #whichList= GoodCopy or BadCopy


        for file in whichList:

            if whichList is GoodCopy: #file name exists as an Access Level ID in spreadsheet
                useAccessListName=file[:-4] #just using the name, no ext
            else: #access list name does NOT exist...
                useAccessListName=file[:-4]

                useAccessListName=known_alternatives(useAccessListName) #looks through dictionary for possible match

                if useAccessListName not in AccessLevels:
                    print(f'{Fore.LIGHTRED_EX}Could not find an alternative for ',str(useAccessListName))
                    time.sleep(0.3)

                    BreakOut=['Pass','pass','Skip','skip']

                    while useAccessListName not in AccessLevels: #Loop user with input requesting valid name for AccessLevel until it checks out
                        useAccessListName=input('The Access List: " '+str(file[:-4])+' " does not exist. Manually enter a known one now: ')
                        if useAccessListName in BreakOut:
                            useAccessListName='SKIP FOR NOW - original file: '+str(file[:-4])
                            break
                    logNewNameUsed(file[:-4],useAccessListName,"(Manual) ")

                else:
                    print(f'{Fore.GREEN}Found alternative for:',str(file[:-4]),f'{Fore.GREEN}using:',str(useAccessListName))
                    time.sleep(0.25)
                    logNewNameUsed(file[:-4],useAccessListName,"(Auto)   ")
                
                print('\n')
    

            collectCardNums=[]
            grabinfo = open(folder_loc+file, "r",encoding='UTF-16')

            
            x=0
            for line in grabinfo:
                x+=1

                if ":" not in line: #only a few instances of this, just outliers and not actual cards
                    if x!=1:
                        Logging(' x Card ID Fail. File: "'+str(file)+'" // (Line '+str(x)+') // Failed line: '+str(line))
                    else:
                        pass #ie 1=the header
                else:

                    getColon=line.index(":")
                        
                    """
                    #if there is no " at index=5, means there is NO first name, ie no comma was used to seperate out names in the csv file
                    
                    if line[5]!='"':
                        firstName=''
                        #get first name now
                        temp1=line[getColon+6:]
                        while temp1[0]==" ":
                            temp1=temp1[1:]
                        lastName=temp1[:(temp1.index(","))]
                    
                    else: #there are both first and last names
                        grabNameToo=line[getColon+6:]
                        while grabNameToo[0]==" ":
                            grabNameToo=grabNameToo[1:]
                        
                        lastName=''
                        while grabNameToo[0]!=",":
                            lastName+=grabNameToo[0]
                            grabNameToo=grabNameToo[1:]

                        grabNameToo=grabNameToo[2:] #comma, and space
                        firstName=''
                        while grabNameToo[0]!='"':
                            firstName+=grabNameToo[0]
                            grabNameToo=grabNameToo[1:]
                    

                    def lastFirst(lastName,firstName): #return combined value for the Cards sheet, Col D
                        if firstName=='':
                            final=str(lastName)
                        else:
                            final=lastName+', '+firstName
                        return final
                    
                    
                    #old way that still exports name, etc info --> no longer used
                    #collectCardNums.append([line[getColon-2:getColon+6],lastName,firstName,lastFirst(lastName,firstName)])
                    """

                    collectCardNums.append([line[getColon-2:getColon+6]]) #card number length is standardized
            
            for elem in collectCardNums: #loop creates rows that will be saved to csv file later
                #old way that used name info, etc
                #arrayToCSV.append([elem[0],useAccessListName,'','',elem[0],elem[1],elem[2],elem[3],''])

                if str(elem[0]) in cardNumbersOnly: #this list is from the first of the merged scripts
                    countGood+=1
                    arrayToCSV.append([elem[0],useAccessListName,''])
                else:
                    countBad+=1

        return arrayToCSV, countGood, countBad


    #looks stupid, but adding the extra space in order to add a comma will format the last line weird
    #this will automatically remove this issue by reading and saving over again
    #I'm sure there is an easier way to do this, but whatever I'm tired and this worked
    def cleanup_csv(location):
        fix=open(location, 'r')
        lines=fix.read().splitlines()

        lines=[line for line in lines if len(line)!=0]
        fix.close()

        fixed=open(location, 'w')
        for line in lines:
            if line!=lines[-1]:
                fixed.writelines(line+"\n")
            else:
                fixed.writelines(line)
        fixed.close()


    #Run for both good and bad copies
    arrayToCSV,countGood,countBad=grabCardInfo(GoodCopy,AccessLevels,arrayToCSV,countGood,countBad) #no issues so far
    arrayToCSV,countGood,countBad=grabCardInfo(BadCopy,AccessLevels,arrayToCSV,countGood,countBad)  #missing some info still on names

    #modify final line to include an extra comma, needed in some cases
    arrayToCSV[-1]=arrayToCSV[-1][:]+['']

    with open(folder_loc+'/result/'+output, 'w') as f:
        write = csv.writer(f)
        write.writerows(arrayToCSV)

    cleanup_csv('Folder/Holding/TheCSVFiles/CSVs/result/'+str(output))

    #Log finished and successful
    Logging(' ++++ Audit report portion of the script executed successfully ++++\n\n')
    Logging(' ---- Program Complete ----\n')

    print(f'{Fore.GREEN}\n\nScript has successfully executed. Look to Folder/Holding/TheCSVFiles/CSVs/results/copyToAuditReportTab.csv for the file.\n\n')


except BaseException as error:

    Logging('\n ! Not a good error : Get this sorted out, program must exit.')

    print(f'{Fore.LIGHTRED_EX}Program must exit from this error. Reference error logs to troubleshoot.')
    waitingOnInput=input(f"{Fore.YELLOW}\nPress [ENTER] to exit the program.")
    sys.exit()


try:
    os.system('start EXCEL.EXE "Folder/Holding/TheCSVFiles/CSVs/result/copyToAuditReportTab.csv"')
except:
    print(f'{Fore.LIGHTRED_EX}Looks like Excel is not installed on this device - no matter, the rest of the script will continue.')
    time.sleep(1)
    Logging(' ! Error: Trying to open Folder/Holding/TheCSVFiles/CSVs/result/copyToAuditReportTab.csv with its default application, which should be Excel.\n')

waitingOnInput=input(f"{Fore.YELLOW}\nPress [ENTER] to exit the program.")
print(f'{Fore.WHITE}Exiting now...')
time.sleep(1.5)

sys.exit()
